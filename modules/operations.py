import os
import random
import shutil
import sys
from pathlib import Path
from typing import Dict, List, Optional, Union

import yaml


def _safe(path: str) -> str:
    """Windows パスのバックスラッシュをスラッシュに統一。"""
    return str(path).replace("\\", "/")


IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")


def get_auto_workers(imgsz: int) -> int:
    """RTX 5060 前提の imgsz に連動した並列数 (workers) 自動決定ロジック。
    - imgsz <= 640  : workers = 4
    - imgsz <= 800  : workers = 2
    - imgsz > 800   : workers = 0 (1024 など)
    """
    if imgsz <= 640:
        return 4
    elif imgsz <= 800:
        return 2
    else:
        return 0



# ---------------------------------------------------------------------------
# 1. データセットフォルダ生成
# ---------------------------------------------------------------------------
def generate_folders(dataset_dir: str) -> List[str]:
    """dataset 配下に train / val フォルダを作成する。"""
    base = Path(dataset_dir)
    folders = [base / "train", base / "val"]
    created = []
    for f in folders:
        f.mkdir(parents=True, exist_ok=True)
        created.append(str(f))
    return created


# ---------------------------------------------------------------------------
# 2. 分割コピー
# ---------------------------------------------------------------------------
def split_copy(
    source_dir: str,
    dataset_dir: str,
    train_count: int,
    val_count: int,
    progress_callback=None,
    cancel_check=None,
) -> dict:
    """元画像からランダム抽出し、train/val フォルダへコピーする。"""
    src = Path(source_dir)
    all_images = [
        p for p in src.iterdir()
        if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS
    ]
    total_needed = train_count + val_count
    if len(all_images) < total_needed:
        raise ValueError(
            f"画像が不足しています。必要: {total_needed} 枚, 利用可能: {len(all_images)} 枚"
        )

    random.shuffle(all_images)
    train_files = all_images[:train_count]
    val_files = all_images[train_count : train_count + val_count]

    train_dst = Path(dataset_dir) / "train"
    val_dst = Path(dataset_dir) / "val"
    train_dst.mkdir(parents=True, exist_ok=True)
    val_dst.mkdir(parents=True, exist_ok=True)

    copied_count = 0
    total_files = len(train_files) + len(val_files)

    for f in train_files:
        if cancel_check and cancel_check():
            raise InterruptedError("処理が中断されました")
        shutil.copy2(str(f), str(train_dst / f.name))
        copied_count += 1
        if progress_callback:
            progress_callback(copied_count, total_files, f.name)

    for f in val_files:
        if cancel_check and cancel_check():
            raise InterruptedError("処理が中断されました")
        shutil.copy2(str(f), str(val_dst / f.name))
        copied_count += 1
        if progress_callback:
            progress_callback(copied_count, total_files, f.name)

    return {"train": len(train_files), "val": len(val_files), "total": total_files}


# ---------------------------------------------------------------------------
# 3. ラベル設定
# ---------------------------------------------------------------------------
def read_text_safe(file_path: Union[str, Path]) -> str:
    """UTF-8, UTF-8-sig, CP932(Shift-JIS) 等のマルチエンコーディング対応で安全にテキストを読み込む。"""
    path = Path(file_path)
    if not path.exists():
        return ""
    for enc in ("utf-8", "utf-8-sig", "cp932", "shift_jis"):
        try:
            return path.read_text(encoding=enc)
        except (UnicodeDecodeError, OSError):
            continue
    try:
        return path.read_bytes().decode("utf-8", errors="ignore")
    except Exception:
        return ""


def save_classes(dataset_dir: str, class_names: List[str]) -> str:
    """classes.txt をデータセット直下および train/val フォルダに保存。
    ルートのパスを返す。"""
    base = Path(dataset_dir)
    content = "\n".join(name.strip() for name in class_names) + "\n"

    # データセット直下
    root_path = base / "classes.txt"
    root_path.parent.mkdir(parents=True, exist_ok=True)
    root_path.write_text(content, encoding="utf-8")

    # train / val サブフォルダ（存在すれば更新、なければ作成）
    for split in ("train", "val"):
        split_dir = base / split
        if split_dir.exists():
            (split_dir / "classes.txt").write_text(content, encoding="utf-8")

    return str(root_path)


# ---------------------------------------------------------------------------
# 4a. labelImg 起動コマンド構築
# ---------------------------------------------------------------------------
def build_labelimg_command(
    python_path: str,
    image_dir: str,
    classes_file: str,
    save_dir: str = None,
) -> List[str]:
    """labelImg 起動用のコマンドリストを返す。"""
    if save_dir is None:
        save_dir = image_dir
    
    # python.exe の場所から labelImg.exe を探す
    py_path = Path(python_path)
    labelimg_exe = py_path.parent / "labelImg.exe"
    
    if labelimg_exe.exists():
        return [
            str(labelimg_exe),
            str(Path(image_dir)),
            str(Path(classes_file)),
            str(Path(save_dir)),
        ]
    else:
        # 見つからない場合は従来の方式を試すが、-c で確実に起動させる
        return [
            str(py_path),
            "-c",
            "import sys; from labelImg.labelImg import main; sys.exit(main())",
            str(Path(image_dir)),
            str(Path(classes_file)),
            str(Path(save_dir)),
        ]


# ---------------------------------------------------------------------------
# YAML 作成
# ---------------------------------------------------------------------------
def create_yaml(
    dataset_dir: str,
    class_names: List[str],
) -> str:
    """data.yaml をデータセット直下に作成。パスを返す。"""
    base = Path(dataset_dir)
    yaml_path = base / "data.yaml"
    names_dict = {i: name.strip() for i, name in enumerate(class_names)}
    content = {
        "path": str(base),
        "train": "train",
        "val": "val",
        "nc": len(class_names),
        "names": names_dict,
    }
    with open(yaml_path, "w", encoding="utf-8") as f:
        yaml.dump(content, f, default_flow_style=False, sort_keys=False, allow_unicode=True)
    return str(yaml_path)


# ---------------------------------------------------------------------------
# キャッシュ削除
# ---------------------------------------------------------------------------
def delete_cache_files(dataset_dir: str) -> int:
    """データセットフォルダ内の .cache ファイルを全て削除。削除数を返す。"""
    base = Path(dataset_dir)
    count = 0
    for cache_file in base.rglob("*.cache"):
        cache_file.unlink()
        count += 1
    return count


# ---------------------------------------------------------------------------
# アノテーションラベル削除 (クリア)
# ---------------------------------------------------------------------------
def clear_annotation_labels(dataset_dir: str) -> int:
    """dataset_dir 配下の train / val フォルダ内の .txt ファイル (classes.txt 除く) を全削除。削除数を返す。"""
    base = Path(dataset_dir)
    count = 0
    for split in ["train", "val"]:
        split_dir = base / split
        if split_dir.exists():
            for txt_file in split_dir.glob("*.txt"):
                if txt_file.name.lower() != "classes.txt":
                    txt_file.unlink(missing_ok=True)
                    count += 1
    return count


# ---------------------------------------------------------------------------
# 5. 学習スクリプト構築
# ---------------------------------------------------------------------------
def build_train_script(
    data_yaml: str,
    epochs: int,
    batch: int,
    imgsz: int,
    workers: int,
    base_model: str,
    yolo_version: str = "11",
    aug_params: dict = None,
) -> str:
    """YOLO 学習用の Python スクリプト文字列を返す。
    freeze_support() 付きで multiprocessing エラーを防止。"""
    if yolo_version == "11":
        model_name = f"yolo11{base_model}.pt"
    else:
        model_name = f"yolov8{base_model}.pt"
    s_data = _safe(data_yaml)

    if aug_params is None:
        aug_params = {}

    hsv_h = aug_params.get("hsv_h", 0.015)
    hsv_s = aug_params.get("hsv_s", 0.7)
    hsv_v = aug_params.get("hsv_v", 0.4)
    degrees = aug_params.get("degrees", 0.0)
    translate = aug_params.get("translate", 0.1)
    scale = aug_params.get("scale", 0.5)
    shear = aug_params.get("shear", 0.0)
    perspective = aug_params.get("perspective", 0.0)
    flipud = aug_params.get("flipud", 0.0)
    fliplr = aug_params.get("fliplr", 0.5)

    return f'''# -*- coding: utf-8 -*-
import multiprocessing
import gc
import torch

multiprocessing.freeze_support()

from ultralytics import YOLO

def main():
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
    gc.collect()

    model = YOLO("{model_name}")
    try:
        results = model.train(
            data="{s_data}",
            epochs={epochs},
            batch={batch},
            imgsz={imgsz},
            workers={workers},
            hsv_h={hsv_h},
            hsv_s={hsv_s},
            hsv_v={hsv_v},
            degrees={degrees},
            translate={translate},
            scale={scale},
            shear={shear},
            perspective={perspective},
            flipud={flipud},
            fliplr={fliplr},
        )
    except Exception as e:
        err_msg = str(e).lower()
        if "out of memory" in err_msg or "cuda" in err_msg:
            print("[WARNING] CUDA Out of Memory (VRAM不足) が発生しました。")
            print("[INFO] VRAMを完全解放し、安全バッチサイズ (batch=4) で自動リトライを開始します...")
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
            gc.collect()
            try:
                results = model.train(
                    data="{s_data}",
                    epochs={epochs},
                    batch=4,
                    imgsz={imgsz},
                    workers={workers},
                    hsv_h={hsv_h},
                    hsv_s={hsv_s},
                    hsv_v={hsv_v},
                    degrees={degrees},
                    translate={translate},
                    scale={scale},
                    shear={shear},
                    perspective={perspective},
                    flipud={flipud},
                    fliplr={fliplr},
                )
            except Exception as e2:
                print("[INFO] VRAMを完全解放し、極小バッチサイズ (batch=2) で最終リトライを開始します...")
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                gc.collect()
                results = model.train(
                    data="{s_data}",
                    epochs={epochs},
                    batch=2,
                    imgsz={imgsz},
                    workers=0,
                    hsv_h={hsv_h},
                    hsv_s={hsv_s},
                    hsv_v={hsv_v},
                    degrees={degrees},
                    translate={translate},
                    scale={scale},
                    shear={shear},
                    perspective={perspective},
                    flipud={flipud},
                    fliplr={fliplr},
                )
        else:
            raise e

    print("Training complete")

if __name__ == "__main__":
    main()
'''


# ---------------------------------------------------------------------------
# 6. 推論→Excel スクリプト文字列生成
# ---------------------------------------------------------------------------
def build_inference_script(
    model_path: str,
    image_dir: str,
    conf: float,
    output_xlsx: str,
) -> str:
    """推論 + Excel 出力を行い、結果を個数別フォルダに保存する。"""
    s_model = _safe(model_path)
    s_images = _safe(image_dir)
    s_xlsx = _safe(output_xlsx)
    return f'''# -*- coding: utf-8 -*-
import multiprocessing
multiprocessing.freeze_support()

import pandas as pd
from ultralytics import YOLO
from pathlib import Path
from tqdm import tqdm
import sys
import shutil

def main():
    model = YOLO("{s_model}")
    exts = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
    image_dir = Path("{s_images}")
    images = [p for p in image_dir.iterdir() if p.suffix.lower() in exts]
    
    # 保存用ベースフォルダ
    save_base = image_dir / "結果画像"
    if save_base.exists():
        shutil.rmtree(save_base)
    save_base.mkdir(parents=True, exist_ok=True)

    print(f"{{len(images)}} 枚を処理中...")
    results = model.predict(source="{s_images}", conf={conf}, save=False,
                            stream=True, verbose=False)
    rows = []
    
    for r in tqdm(results, total=len(images), desc="推論進捗", unit="img", file=sys.stdout):
        fname = Path(r.path).name
        count = len(r.boxes)
        
        # 個数別フォルダの作成と保存
        count_dir = save_base / str(count)
        count_dir.mkdir(parents=True, exist_ok=True)
        save_path = count_dir / fname
        r.save(filename=str(save_path))

        if count > 0:
            for box in r.boxes:
                cls_id = int(box.cls[0])
                cls_name = model.names[cls_id]
                confidence = float(box.conf[0])
                rows.append({{
                    "ファイル名": fname,
                    "クラス名": cls_name,
                    "信頼度": round(confidence, 4),
                    "検出個数": count
                }})
        else:
            rows.append({{"ファイル名": fname, "クラス名": "None", "信頼度": 0, "検出個数": 0}})

    df = pd.DataFrame(rows)
    df.to_excel("{s_xlsx}", index=False, engine="openpyxl")
    
    # オートフィルター設定
    try:
        from openpyxl import load_workbook
        wb = load_workbook("{s_xlsx}")
        ws = wb.active
        ws.auto_filter.ref = ws.dimensions
        wb.save("{s_xlsx}")
    except:
        pass

    print(f"\\nExcel出力完了: {s_xlsx}")
    print(f"結果画像保存先: {{save_base}}")
    print(f"処理画像数: {{len(images)}}枚, 検出総数: {{len(rows)}}個")

if __name__ == "__main__":
    main()
'''


# ---------------------------------------------------------------------------
# 4d. ラベル一覧 Excel 出力
# ---------------------------------------------------------------------------
def export_label_list_excel(dataset_dir: str) -> str:
    """dataset フォルダ内に ラベル一覧.xlsx を出力し、その絶対パスを返す。"""
    base = Path(dataset_dir)
    xlsx_path = base / "ラベル一覧.xlsx"

    classes_map = {}
    cls_file = base / "classes.txt"
    if cls_file.exists():
        lines = cls_file.read_text(encoding="utf-8").splitlines()
        for idx, line in enumerate(lines):
            if line.strip():
                classes_map[idx] = line.strip()

    rows = []
    exts = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")

    for split in ["train", "val"]:
        split_dir = base / split
        if not split_dir.exists():
            continue

        images = [p for p in split_dir.iterdir() if p.suffix.lower() in exts]
        for img in images:
            txt_file = split_dir / (img.stem + ".txt")
            has_txt = "あり" if txt_file.exists() else "なし"
            count = 0
            detected_classes = []

            if txt_file.exists():
                txt_lines = txt_file.read_text(encoding="utf-8").splitlines()
                for line in txt_lines:
                    parts = line.strip().split()
                    if parts:
                        count += 1
                        try:
                            c_id = int(parts[0])
                            c_name = classes_map.get(c_id, f"ID:{c_id}")
                            if c_name not in detected_classes:
                                detected_classes.append(c_name)
                        except ValueError:
                            pass

            cls_str = ", ".join(detected_classes) if detected_classes else "-"
            rows.append({
                "データ種別": split,
                "画像ファイル名": img.name,
                "アノテーション": has_txt,
                "ラベル数": count,
                "検出クラス": cls_str
            })

    import pandas as pd
    df = pd.DataFrame(rows)
    df.to_excel(str(xlsx_path), index=False, engine="openpyxl")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_path))
        ws = wb.active
        ws.auto_filter.ref = ws.dimensions
        wb.save(str(xlsx_path))
    except Exception:
        pass

    return str(xlsx_path)


# ---------------------------------------------------------------------------
# 4c. 自動アノテーション用スクリプト文字列生成
#     verbose=False で個別結果を非表示、tqdm のみ進捗表示
# ---------------------------------------------------------------------------
def build_autolabel_script(
    model_path: str,
    dataset_dir: str,
    conf: float,
    class_names: List[str],
) -> str:
    """自動アノテーション用の Python スクリプト文字列を返す。"""
    s_model = _safe(model_path)
    s_dataset = _safe(dataset_dir)
    return f'''# -*- coding: utf-8 -*-
import multiprocessing
multiprocessing.freeze_support()

import os
import sys
import yaml
import pandas as pd
from ultralytics import YOLO
from pathlib import Path
from tqdm import tqdm

def main():
    model = YOLO("{s_model}")
    dataset = Path("{s_dataset}")

    splits = []
    for split in ["train", "val"]:
        d = dataset / split
        if d.exists():
            splits.append((split, d))

    if not splits:
        print("ERROR: train または val フォルダが見つかりません")
        sys.exit(1)

    # classes.txt を各フォルダに作成
    names_list = [model.names[i] for i in range(len(model.names))]
    for split, split_dir in splits:
        cls_file = split_dir / "classes.txt"
        with open(cls_file, "w", encoding="utf-8") as f:
            for name in names_list:
                f.write(name + "\\n")

    excel_rows = []
    total_images = 0

    for split, split_dir in splits:
        exts = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
        images = [p for p in split_dir.iterdir() if p.suffix.lower() in exts]
        total_images += len(images)

        print(f"\\n[{{split}}] {{len(images)}} 枚を処理中...")
        results = model.predict(source=str(split_dir), conf={conf},
                                save=False, stream=True, verbose=False)

        for r in tqdm(results, total=len(images), desc=f"  {{split}}",
                      unit="img", file=sys.stdout):
            img_path = Path(r.path)
            txt_path = split_dir / (img_path.stem + ".txt")
            fname = img_path.name
            count = len(r.boxes)

            if count > 0:
                # save_txt はデフォルトで上書きするが、明示的に。
                r.save_txt(str(txt_path))
                for box in r.boxes:
                    cls_id = int(box.cls[0])
                    cls_name = model.names[cls_id]
                    conf_val = float(box.conf[0])
                    excel_rows.append({{
                        "データ種別": split,
                        "画像ファイル名": fname,
                        "クラス名": cls_name,
                        "信頼度": round(conf_val, 4),
                        "検出個数": count
                    }})
            else:
                # 検出なしの場合、ファイルを空（上書き）にして確実に既存ラベルを消す
                txt_path.write_text("", encoding="utf-8")
                excel_rows.append({{
                    "データ種別": split,
                    "画像ファイル名": fname,
                    "クラス名": "None",
                    "信頼度": 0,
                    "検出個数": 0
                }})

    xlsx_path = dataset / "autolabel_result.xlsx"
    df = pd.DataFrame(excel_rows)
    df.to_excel(str(xlsx_path), index=False, engine="openpyxl")

    # オートフィルター設定
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_path))
        ws = wb.active
        ws.auto_filter.ref = ws.dimensions
        wb.save(str(xlsx_path))
    except:
        pass

    print(f"\\n=== オートラベリング完了 ===")
    print(f"  処理画像数: {{total_images}}")
    print(f"  Excel出力 : {{xlsx_path}}")

if __name__ == "__main__":
    main()
'''


# ---------------------------------------------------------------------------
# 7. NCNN エクスポートスクリプト
# ---------------------------------------------------------------------------
def build_ncnn_export_script(model_path: str, imgsz: int) -> str:
    """NCNN 形式にモデルをエクスポートし、作成されたフォルダの末尾に .ncnn を付与するスクリプト文字列を返す。"""
    s_model = _safe(model_path)
    return f'''# -*- coding: utf-8 -*-
import multiprocessing
multiprocessing.freeze_support()

import os
import shutil
from pathlib import Path
from ultralytics import YOLO

def main():
    model = YOLO("{s_model}")
    out = model.export(format="ncnn", imgsz={imgsz})
    if out:
        out_path = Path(out)
        if out_path.exists():
            if not out_path.name.endswith(".ncnn"):
                new_path = out_path.with_name(out_path.name + ".ncnn")
                if new_path.exists():
                    if new_path.is_dir():
                        shutil.rmtree(new_path)
                    else:
                        new_path.unlink()
                out_path.rename(new_path)
                print(f"NCNN model saved to: {{new_path}}")
            else:
                print(f"NCNN model saved to: {{out_path}}")

if __name__ == "__main__":
    main()
'''


# ---------------------------------------------------------------------------
# 8. 未アノテーション画像に対する空ラベルファイルの自動作成
# ---------------------------------------------------------------------------
def create_empty_labels(dataset_dir: Union[str, Path]) -> int:
    """train / val フォルダ内の未アノテーション画像（対応する .txt が存在しない画像）に対して空の .txt ファイルを作成する。
    戻り値: 作成した空ラベルファイル数
    """
    base = Path(dataset_dir)
    if not base.exists():
        return 0

    created_count = 0
    # 通常の dataset/{train,val} または dataset/images/{train,val} の両方に対応
    target_dirs = []
    for split in ["train", "val"]:
        # 1. dataset/train, dataset/val
        d1 = base / split
        if d1.is_dir():
            target_dirs.append((d1, d1))
        # 2. dataset/images/train, dataset/labels/train
        img_d = base / "images" / split
        lbl_d = base / "labels" / split
        if img_d.is_dir():
            lbl_d.mkdir(parents=True, exist_ok=True)
            target_dirs.append((img_d, lbl_d))

    for img_dir, lbl_dir in target_dirs:
        for img_file in img_dir.iterdir():
            if img_file.is_file() and img_file.suffix.lower() in IMAGE_EXTENSIONS:
                txt_file = lbl_dir / f"{img_file.stem}.txt"
                if not txt_file.exists():
                    txt_file.touch()  # 空ファイル作成
                    created_count += 1

    return created_count


# ---------------------------------------------------------------------------
# 9. デスクトップ起動ショートカットの作成 (OS自動判定)
# ---------------------------------------------------------------------------
def create_desktop_shortcut(python_path: str = "") -> dict:
    """デスクトップ上に 'YOLOマネージャー' の起動ショートカット/スクリプトを作成する。
    - Windows: 'YOLOマネージャー.bat'
    - Linux / Raspberry Pi: 'YOLOマネージャー.sh' (chmod 0o755) & 'YOLOマネージャー.desktop'
    """
    app_dir = Path(__file__).resolve().parent.parent
    main_py = app_dir / "modules" / "main.py"
    
    # 実行用 python パス
    py_exec = python_path if python_path and Path(python_path).exists() else sys.executable

    # デスクトップのパスを取得
    home = Path.home()
    desktop_dir = home / "Desktop"
    if not desktop_dir.exists():
        desktop_jp = home / "デスクトップ"
        if desktop_jp.exists():
            desktop_dir = desktop_jp
        else:
            desktop_dir.mkdir(parents=True, exist_ok=True)

    is_windows = sys.platform == "win32"

    if is_windows:
        bat_file = desktop_dir / "YOLOマネージャー.bat"
        content = (
            f"@echo off\n"
            f"chcp 65001 > nul\n"
            f"cd /d \"{app_dir}\"\n"
            f"\"{py_exec}\" \"{main_py}\"\n"
            f"if errorlevel 1 pause\n"
        )
        bat_file.write_text(content, encoding="utf-8")
        return {
            "success": True,
            "path": str(bat_file),
            "type": "bat",
            "message": f"Windows用実行ファイル ({bat_file.name}) をデスクトップに作成しました。",
        }
    else:
        # Linux / Raspberry Pi (.sh のみ作成、ターミナルウィンドウを表示)
        sh_file = desktop_dir / "YOLOマネージャー.sh"
        sh_content = (
            f"#!/bin/bash\n"
            f"# ターミナル環境でない場合（デスクトップダブルクリックなど）はターミナルを起動して自身を実行\n"
            f"if [ ! -t 0 ] && [ -z \"$ALREADY_IN_TERMINAL\" ]; then\n"
            f"    export ALREADY_IN_TERMINAL=1\n"
            f"    for term in lxterminal x-terminal-emulator gnome-terminal xterm konsole; do\n"
            f"        if command -v $term >/dev/null 2>&1; then\n"
            f"            exec $term -e \"$0\" \"$@\"\n"
            f"        fi\n"
            f"    done\n"
            f"fi\n\n"
            f"cd \"{app_dir}\"\n"
            f"\"{py_exec}\" \"{main_py}\"\n\n"
            f"echo ''\n"
            f"read -p \"処理が終了しました。Enterキーを押すと閉じます...\"\n"
        )
        sh_file.write_text(sh_content, encoding="utf-8")
        try:
            sh_file.chmod(0o755)
        except Exception:
            pass

        return {
            "success": True,
            "path": str(sh_file),
            "type": "sh",
            "message": f"Linux/Raspberry Pi用実行スクリプト ({sh_file.name}) をデスクトップに作成しました。",
        }


# ---------------------------------------------------------------------------
# 10. Raspberry Pi 環境判定
# ---------------------------------------------------------------------------
def is_raspberry_pi() -> bool:
    """現在の実行環境が Raspberry Pi (または ARM SBC) かどうかを判定する。"""
    try:
        model_path = Path("/proc/device-tree/model")
        if model_path.exists():
            model_text = model_path.read_text(encoding="utf-8", errors="ignore").lower()
            if "raspberry pi" in model_text or "rpi" in model_text:
                return True
    except Exception:
        pass

    try:
        for p in ["/etc/rpi-issue", "/etc/os-release"]:
            f = Path(p)
            if f.exists():
                content = f.read_text(encoding="utf-8", errors="ignore").lower()
                if "raspberry pi" in content or "raspbian" in content:
                    return True
    except Exception:
        pass

    if sys.platform.startswith("linux"):
        import platform
        mach = platform.machine().lower()
        if "arm" in mach or "aarch64" in mach:
            return True

    return False


